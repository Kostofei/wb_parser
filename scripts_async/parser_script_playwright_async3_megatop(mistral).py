import os
import re
import asyncio
import pandas as pd
from functools import wraps
from openpyxl import Workbook
from openpyxl.styles import Alignment
import aiofiles
from io import BytesIO
from asyncio import to_thread
from colorama import Fore, Style, init
from playwright.async_api import (
    async_playwright,
    TimeoutError as PlaywrightTimeoutError,
    Page,
    Playwright,
    Browser,
    BrowserContext,
)

init()

# Constants
TARGET_URL = "https://www.wildberries.by"
EXCLUDED_CATEGORIES = ['бренды', 'wibes', 'экспресс', 'акции', 'грузовая доставка']
TIME_WAIT = 1000

def timeit(func):
    @wraps(func)
    async def async_wrapper(*args, **kwargs):
        start_time = asyncio.get_event_loop().time()
        try:
            result = await func(*args, **kwargs)
            end_time = asyncio.get_event_loop().time()
            print(f'Function {func.__name__} executed in {end_time - start_time:.4f} sec.')
            return result
        except Exception as e:
            end_time = asyncio.get_event_loop().time()
            print(f'Function {func.__name__} failed after {end_time - start_time:.4f} sec.')
            raise e

    def wrapper(*args, **kwargs):
        if asyncio.iscoroutinefunction(func):
            return async_wrapper(*args, **kwargs)
        else:
            start_time = asyncio.get_event_loop().time()
            result = func(*args, **kwargs)
            end_time = asyncio.get_event_loop().time()
            print(f'Synchronous function {func.__name__} executed in {end_time - start_time:.4f} sec.')
            return result

    return wrapper

@timeit
async def run_wb_parser():
    result = await parse_all_categories()
    if result:
        await process_categories_to_excel(result)

async def process_categories_to_excel(
        initial_data: list[dict],
        output_filename: str = "categories.xlsx",
        exclude_root_in_path: bool = True
) -> None:
    unique_filename = await get_unique_filename(output_filename)
    await _write_categories_to_excel(initial_data, unique_filename, exclude_root_in_path)

async def get_unique_filename(base_filename: str) -> str:
    if not await to_thread(os.path.exists, base_filename):
        return base_filename

    name, ext = os.path.splitext(base_filename)
    directory = os.path.dirname(base_filename) or '.'
    existing_files = await to_thread(
        lambda: [f for f in os.listdir(directory) if f.startswith(name) and f.endswith(ext)]
    )

    if not existing_files:
        return base_filename

    max_suffix = max(
        (int(re.search(r'_(\d+)$', f[len(name):-len(ext)]).group(1)) if re.search(r'_(\d+)$', f[len(name):-len(ext)]) else 0 for f in existing_files),
        default=0
    )

    return f"{name}_{max_suffix + 1}{ext}"

async def _write_categories_to_excel(
        initial_data: list[dict],
        output_filename: str,
        exclude_root_in_path: bool
) -> None:
    def _sanitize_sheet_name(name: str) -> str:
        sanitized_name = re.sub(r'[\\/*?:\[\]]', '', name)
        return sanitized_name[:31]

    def _get_categories_with_levels(
            item: dict,
            current_level: int = 0,
            parent_path: list[str] | None = None,
            last_category: str | None = None
    ) -> list[dict]:
        categories = []
        current_path = parent_path.copy() if parent_path else []
        if not exclude_root_in_path or current_level > 0:
            current_path.append(item['name'])

        item_categories = item.get('Категория', [])
        if isinstance(item_categories, list):
            for category in item_categories:
                if category == "Категорий нет" and last_category:
                    modified_path = current_path[:-1] if current_path else []
                    categories.append({
                        'Категория': last_category,
                        'Уровень': current_level,
                        'Путь': ' → '.join(modified_path) if modified_path else item['name']
                    })
                else:
                    categories.append({
                        'Категория': category,
                        'Уровень': current_level,
                        'Путь': ' → '.join(current_path) if current_path else item['name']
                    })
                if category != "Категорий нет":
                    last_category = category
        else:
            if item_categories == "Категорий нет" and last_category:
                modified_path = current_path[:-1] if current_path else []
                categories.append({
                    'Категория': last_category,
                    'Уровень': current_level,
                    'Путь': ' → '.join(modified_path) if modified_path else item['name']
                })
            else:
                categories.append({
                    'Категория': item_categories,
                    'Уровень': current_level,
                    'Путь': ' → '.join(current_path) if current_path else item['name']
                })
            if item_categories != "Категорий нет":
                last_category = item_categories

        if 'subcategories' in item:
            for sub in item['subcategories']:
                categories.extend(_get_categories_with_levels(
                    sub, current_level + 1, current_path.copy(), last_category
                ))

        return categories

    wb = Workbook()
    wb.remove(wb.active)

    for section in initial_data:
        section_id = section.get('data_menu_id', 'N/A')
        section_name = _sanitize_sheet_name(f"{section['name']} (id {section_id})")
        all_categories = []

        if 'subcategories' in section:
            for subcategory in section['subcategories']:
                all_categories.extend(_get_categories_with_levels(
                    subcategory, current_level=1, parent_path=[], last_category=None
                ))
        elif 'Категория' in section:
            all_categories.extend(_get_categories_with_levels(
                section, current_level=0, parent_path=[], last_category=None
            ))

        df = pd.DataFrame(all_categories)
        if not df.empty:
            df = df[['Категория', 'Уровень', 'Путь']]
            ws = wb.create_sheet(title=section_name)
            ws.append(['Категория', 'Уровень', 'Путь'])

            for _, row in df.iterrows():
                ws.append([row['Категория'], row['Уровень'], row['Путь']])

            for row in ws.iter_rows(min_row=2, min_col=2, max_col=2):
                for cell in row:
                    cell.alignment = Alignment(horizontal='center', vertical='center')

            ws['B1'].alignment = Alignment(horizontal='center', vertical='center')
            ws.column_dimensions['A'].width = 30
            ws.column_dimensions['B'].width = 10
            ws.column_dimensions['C'].width = 50
        else:
            wb.create_sheet(title=section_name)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    async with aiofiles.open(output_filename, 'wb') as f:
        await f.write(buffer.getvalue())

async def create_browser_session(p: Playwright) -> tuple[Browser, BrowserContext, Page]:
    browser = await p.chromium.launch(
        headless=True,
        timeout=60000,
        args=[
            '--disable-blink-features=AutomationControlled',
            '--disable-infobars',
            '--start-maximized',
            '--no-sandbox',
            '--disable-setuid-sandbox',
            '--disable-dev-shm-usage',
            '--disable-web-security',
            '--disable-notifications'
        ]
    )

    context = await browser.new_context(
        viewport={'width': 1280, 'height': 720},
        user_agent='Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36',
        bypass_csp=True,
        java_script_enabled=True
    )

    page = await context.new_page()
    await page.route("**/*", route_handler)

    return browser, context, page

async def route_handler(route, request):
    blocked_types = ["image", "media", "other"]
    if request.resource_type in blocked_types:
        await route.abort()
    else:
        await route.continue_()

@timeit
async def load_main_categories(page: Page) -> list:
    result = []

    try:
        await page.goto(TARGET_URL, timeout=120000, wait_until="networkidle")
        button_burger = page.locator('button.nav-element__burger.j-menu-burger-btn').first
        if button_burger:
            await button_burger.click()

        await page.wait_for_selector('ul.menu-burger__main-list a.menu-burger__main-list-link:has-text("Бренды"):visible')
        main_categories = await page.locator('ul.menu-burger__main-list > li.menu-burger__main-list-item').all()

        for category in main_categories:
            category_link = category.locator('a.menu-burger__main-list-link')
            if not category_link:
                continue

            name_category = (await category_link.inner_text()).strip()
            if name_category.lower() in EXCLUDED_CATEGORIES:
                continue

            result.append({
                'name': name_category,
                'url': await category_link.get_attribute('href'),
                'data_menu_id': await category.get_attribute('data-menu-id'),
                'subcategories': []
            })
    except PlaywrightTimeoutError:
        print("Timeout waiting for element")
    except Exception as e:
        print(f"Error occurred: {str(e)}")

    return result

async def load_subcategories(
        category: dict,
        page: Page,
        sem: asyncio.Semaphore,
        level: int = 1,
        flag: bool = True,
        max_retries: int = 3
) -> dict:
    async with sem:
        if flag:
            print(f"{level * '-' + ' ' if level != 1 else ''}{category['name']}")

        for attempt in range(max_retries):
            try:
                await page.goto(f"{TARGET_URL}{category['url']}", timeout=120000, wait_until="domcontentloaded")

                # Wait a bit to ensure the page is fully loaded
                await asyncio.sleep(2)

                menu_subcategory = page.locator('ul.menu-category__subcategory')
                if await menu_subcategory.count() > 0:
                    await process_menu_items(page, category, 'subcategory', level)
                    break
                else:
                    menu_list = page.locator('ul.menu-category__list')
                    if await menu_list.count() > 0:
                        await process_menu_items(page, category, 'list', level)
                        break
                    else:
                        category_filter = page.locator("div.dropdown-filter:has-text('Категория'):visible")
                        if await category_filter.count() > 0:
                            await category_filter.hover()
                            await load_and_collect_categories(page, category, level)
                            break
                        else:
                            btm_burger = page.locator('button.dropdown-filter__btn--burger > div.dropdown-filter__btn-name').first
                            if await btm_burger.text_content() not in category.get('parent', []):
                                await btm_burger.hover()
                                await load_and_collect_subcategories(page, category, level)
                                break
                            else:
                                category['Категория'] = 'Категорий нет'
                                break
            except Exception as e:
                print(f"Attempt {attempt + 1} failed for {category['name']}: {str(e)}")
                if attempt == max_retries - 1:
                    print(f"Max retries reached for {category['name']}")
                    category['Категория'] = 'Error loading category'
                await asyncio.sleep(2)  # Wait before retrying

    return category

async def process_menu_items(
        page: Page,
        category: dict,
        menu_type: str,
        level: int
) -> None:
    selector = {
        'subcategory': ('li.menu-category__subcategory-item', 'a.menu-category__subcategory-link'),
        'list': ('li.menu-category__item', 'a.menu-category__link')
    }[menu_type]

    result = []
    menu_items = await page.locator(selector[0]).all()

    for item in menu_items:
        if menu_type == 'list' and await item.locator('p.menu-category__item').count() > 0:
            continue

        link = item.locator(selector[1])
        if await link.count() > 0:
            parent = category['parent'].copy() if category.get('parent') else []
            parent.append(category['name'])
            result.append({
                'name': (await link.inner_text()).strip(),
                'url': await link.get_attribute('href'),
                'parent': parent,
                'level': level,
            })

    await print_results_and_load_subcategories(page, category, result, level)

async def print_results_and_load_subcategories(
        page: Page,
        category: dict,
        result: list,
        level: int,
        max_concurrent_tasks: int = 2
) -> None:
    category['subcategories'] = []

    if result:
        sub_sem = asyncio.Semaphore(max_concurrent_tasks)
        tasks = [load_subcategories(item, page, sub_sem, level + 1, False) for item in result]
        category['subcategories'] = await asyncio.gather(*tasks, return_exceptions=True)

async def load_and_collect_categories(
        page: Page,
        category: dict,
        level: int
) -> None:
    await page.wait_for_timeout(100)
    show_all_button = page.locator("button.filter__show-all:has-text('Показать все'):visible").first

    if show_all_button and await show_all_button.is_visible():
        await show_all_button.click()
        current_items = await page.locator('li.filter__item:visible').all()

        while True:
            if not current_items:
                await page.wait_for_timeout(500)
                current_items = await page.locator('li.filter__item:visible').all()
                continue

            await current_items[-1].hover()
            await page.wait_for_timeout(100)
            new_items = await page.locator('li.filter__item:visible').all()

            if len(new_items) == len(current_items):
                break

            current_items = new_items
    else:
        current_items = await page.locator('li.filter__item:visible').all()

    result = []
    for item in current_items:
        link = item.locator('span.checkbox-with-text__text')
        if link:
            result.append((await link.inner_text()).strip())

    category['Категория'] = result

async def load_and_collect_subcategories(
        page: Page,
        category: dict,
        level: int,
        max_concurrent_tasks: int = 2
) -> None:
    await page.wait_for_timeout(100)
    all_items = await page.locator('ul.filter-category__list > li.filter-category__item').all()
    result = []

    for item in all_items:
        link = item.locator('a.filter-category__link').first
        if link:
            parent = category['parent'].copy() if category.get('parent') else []
            parent.append(category['name'])
            result.append({
                'name': (await link.inner_text()).strip(),
                'url': await link.get_attribute('href'),
                'parent': parent,
                'level': level,
            })

    if result:
        sub_sem = asyncio.Semaphore(max_concurrent_tasks)
        tasks = [load_subcategories(item, page, sub_sem, level + 1, False) for item in result]
        category['subcategories'] = await asyncio.gather(*tasks, return_exceptions=True)

async def parse_all_categories(max_concurrent_tasks: int = 3) -> list | None:
    async with async_playwright() as p:
        browser, context, page = await create_browser_session(p)

        try:
            print('Получаю категории')
            categories = await load_main_categories(page)

            if not categories:
                raise ValueError("Не удалось загрузить категории (пустой список)")

            print('- Получаю подкатегории для категорий')
            sem = asyncio.Semaphore(max_concurrent_tasks)
            tasks = [load_subcategories(category, page, sem) for category in categories[:5]]
            results = await asyncio.gather(*tasks, return_exceptions=True)

            valid_results = [r for r in results if not isinstance(r, Exception)]
            return valid_results
        except PlaywrightTimeoutError as e:
            print(f"Timeout error occurred: {e}")
            return None
        except Exception as e:
            print(f"Other error occurred: {e}")
            return None
        finally:
            await page.close()
            await context.close()
            await browser.close()

if __name__ == "__main__":
    asyncio.run(run_wb_parser())
