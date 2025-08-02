# Обновленный код с использованием asyncio.Queue и одной переиспользуемой вкладки

import os
import re
import time
import asyncio
import pandas as pd
from functools import wraps
from openpyxl import Workbook
from openpyxl.styles import Alignment
import aiofiles
from io import BytesIO
from asyncio import to_thread
from colorama import Fore, Style, init
from playwright.async_api import async_playwright, Page, Playwright, Browser, BrowserContext

init()

TARGET_URL = "https://www.wildberries.by"
EXCLUDED_CATEGORIES = ['бренды', 'wibes', 'экспресс', 'акции', 'грузовая доставка']
TIME_WAIT = 1000

def timeit(func):
    @wraps(func)
    async def async_wrapper(*args, **kwargs):
        start_time = time.time()
        try:
            result = await func(*args, **kwargs)
            end_time = time.time()
            print(f'Функция {func.__name__} выполнилась за {end_time - start_time:.4f} сек.')
            return result
        except Exception as e:
            end_time = time.time()
            print(f'Функция {func.__name__} завершилась с ошибкой за {end_time - start_time:.4f} сек.')
            raise e
    return async_wrapper


class SharedPage:
    def __init__(self, context: BrowserContext):
        self.context = context
        self.page: Page | None = None
        self.lock = asyncio.Lock()

    async def init(self):
        self.page = await self.context.new_page()
        await self.page.route("**/*", self.route_handler)

    async def route_handler(self, route, request):
        blocked_types = ["image", "media", "other"]
        if request.resource_type in blocked_types:
            await route.abort()
        else:
            await route.continue_()

    async def use(self, coro):
        async with self.lock:
            return await coro(self.page)

    async def close(self):
        if self.page:
            await self.page.close()


async def create_browser_session(p: Playwright) -> tuple[Browser, BrowserContext]:
    browser = await p.chromium.launch(headless=True)
    context = await browser.new_context(
        viewport={"width": 1280, "height": 900},
        user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                   "(KHTML, like Gecko) Chrome/114.0.0.0 Safari/537.36"
    )
    return browser, context


async def load_main_categories(context: BrowserContext):
    page = await context.new_page()
    await page.goto(TARGET_URL, timeout=60000)
    await page.wait_for_timeout(1000)
    menu_items = await page.locator('li.menu-burger__main-list-item').all()
    result = []
    for item in menu_items:
        link = item.locator('a.menu-burger__main-list-link')
        if await link.count() > 0:
            text = (await link.inner_text()).strip().lower()
            if text not in EXCLUDED_CATEGORIES:
                result.append({
                    'name': (await link.inner_text()).strip(),
                    'url': await link.get_attribute('href'),
                    'parent': [],
                    'level': 1
                })
    await page.close()
    return result


@timeit
async def run_wb_parser():
    result = await parse_all_categories()
    await process_categories_to_excel(result)


async def parse_all_categories():
    async with async_playwright() as p:
        browser, context = await create_browser_session(p)
        shared_page = SharedPage(context)
        await shared_page.init()

        try:
            categories = await load_main_categories(context)
            queue = asyncio.Queue()
            for cat in categories:
                await queue.put(cat)

            tasks = [asyncio.create_task(queue_worker(queue, shared_page)) for _ in range(3)]
            await queue.join()
            for _ in tasks:
                await queue.put(None)
            await asyncio.gather(*tasks)
            return categories
        finally:
            await shared_page.close()
            await context.close()
            await browser.close()


async def queue_worker(queue: asyncio.Queue, shared_page: SharedPage):
    while True:
        category = await queue.get()
        if category is None:
            queue.task_done()
            break
        await process_category(category, shared_page, queue)
        queue.task_done()


async def process_category(category: dict, shared_page: SharedPage, queue: asyncio.Queue, level: int = 1):
    category_parent = f'. {Fore.CYAN}Родитель - {category.get("parent")}{Style.RESET_ALL}' if category.get("parent") else ""
    print(f'{level * "-" + " " if level != 1 else ""}{category["name"]}{category_parent}')

    async def _task(page: Page):
        try:
            await page.goto(f"{TARGET_URL}{category['url']}", timeout=120000, wait_until="domcontentloaded")
            await page.wait_for_timeout(300)

            menu_subcategory = page.locator('ul.menu-category__subcategory')
            if await menu_subcategory.count() > 0:
                await process_menu_items(page, category, 'subcategory', level, queue)
                return

            menu_list = page.locator('ul.menu-category__list')
            if await menu_list.count() > 0:
                await process_menu_items(page, category, 'list', level, queue)
                return

            category_filter = page.locator("div.dropdown-filter:has-text('Категория')")
            if await category_filter.count() > 0:
                await category_filter.hover()
                await load_and_collect_categories(page, category, level)
                return

            btm_burger = page.locator('button.dropdown-filter__btn--burger > div.dropdown-filter__btn-name').first
            if await btm_burger.count() and (await btm_burger.text_content()) not in category.get('parent', ''):
                await btm_burger.hover()
                await load_and_collect_subcategories(page, category, level, queue)
                return
            else:
                category['Категория'] = 'Категорий нет'
        except Exception as e:
            print(f"{Fore.RED}Обработка ошибок {category['name']}: {str(e)}{Style.RESET_ALL}")

    await shared_page.use(_task)


async def process_menu_items(page: Page, category: dict, menu_type: str, level: int, queue: asyncio.Queue):
    selector = {
        'subcategory': ('li.menu-category__subcategory-item', 'a.menu-category__subcategory-link'),
        'list': ('li.menu-category__item', 'a.menu-category__link')
    }[menu_type]
    result = []
    menu_items = await page.locator(selector[0]).all()
    for item in menu_items:
        if menu_type == 'list' and await item.locator('p.menu-category__item').count() > 0:
            continue
        link = item.locator(selector[1])
        if await link.count() > 0:
            parent = category['parent'].copy() if category.get('parent') else []
            parent.append(category['name'])
            result.append({
                'name': (await link.inner_text()).strip(),
                'url': await link.get_attribute('href'),
                'parent': parent,
                'level': level
            })
    category['subcategories'] = result
    for subcat in result:
        await queue.put(subcat)


async def load_and_collect_categories(page: Page, category: dict, level: int):
    await page.wait_for_timeout(100)
    show_all_button = page.locator("button.filter__show-all:has-text('Показать все')").first
    if show_all_button and await show_all_button.is_visible():
        await show_all_button.click()
        current_items = await page.locator('li.filter__item:visible').all()
        while True:
            if not current_items:
                await page.wait_for_timeout(500)
                current_items = await page.locator('li.filter__item:visible').all()
                continue
            await current_items[-1].hover()
            await page.wait_for_timeout(100)
            new_items = await page.locator('li.filter__item:visible').all()
            if len(new_items) == len(current_items):
                break
            current_items = new_items
    else:
        current_items = await page.locator('li.filter__item:visible').all()

    result = []
    for item in current_items:
        link = item.locator('span.checkbox-with-text__text')
        if link:
            result.append((await link.inner_text()).strip())
    category['Категория'] = result


async def load_and_collect_subcategories(page: Page, category: dict, level: int, queue: asyncio.Queue):
    result = []
    await page.wait_for_timeout(100)
    all_items = await page.locator('ul.filter-category__list > li.filter-category__item').all()
    for item in all_items:
        link = item.locator('a.filter-category__link').first
        if link:
            parent = category['parent'].copy() if category.get('parent') else []
            parent.append(category['name'])
            result.append({
                'name': (await link.inner_text()).strip(),
                'url': await link.get_attribute('href'),
                'parent': parent,
                'level': level
            })
    category['subcategories'] = result
    for subcat in result:
        await queue.put(subcat)


async def process_categories_to_excel(
        initial_data: list[dict],
        output_filename: str = "categories.xlsx",
        exclude_root_in_path: bool = True
) -> None:
    """
    Обрабатывает категории и сохраняет их в Excel файл с уровнями вложенности
    Args:
        initial_data: Исходные данные категорий (список словарей)
        output_filename: Имя выходного файла Excel
        exclude_root_in_path: Исключать ли корневой элемент из пути
    """
    # Получаем уникальное имя файла
    unique_filename = await get_unique_filename(output_filename)

    # Вызываем функцию для записи данных в файл
    await _write_categories_to_excel(initial_data, unique_filename, exclude_root_in_path)


async def get_unique_filename(base_filename: str) -> str:
    """
    Асинхронно генерирует уникальное имя файла, если файл с таким именем уже существует.
    Args:
        base_filename: Базовое имя файла
    Returns:
        Уникальное имя файла
    """
    # Асинхронно проверяем существование файла
    if not await to_thread(os.path.exists, base_filename):
        return base_filename

    # Разделяем имя файла и расширение
    name, ext = os.path.splitext(base_filename)
    directory = os.path.dirname(base_filename) or '.'

    # Асинхронно получаем список файлов в директории
    existing_files = await to_thread(
        lambda: [f for f in os.listdir(directory)
                 if f.startswith(name) and f.endswith(ext)]
    )

    if not existing_files:
        return base_filename

    # Находим максимальный суффикс
    max_suffix = 0
    for f in existing_files:
        # Пытаемся извлечь суффикс из имени файла
        try:
            suffix_part = f[len(name):-len(ext)]
            # Обрабатываем случай с подчеркиванием (например, "file_1.xlsx")
            if '_' in suffix_part:
                suffix = int(suffix_part.split('_')[-1])
            else:
                suffix = int(suffix_part)
            if suffix > max_suffix:
                max_suffix = suffix
        except ValueError:
            continue

    # Генерируем новое имя файла
    new_filename = f"{name}_{max_suffix + 1}{ext}" if max_suffix > 0 else f"{name}_1{ext}"
    return new_filename


async def _write_categories_to_excel(
        initial_data: list[dict],
        output_filename: str,
        exclude_root_in_path: bool
) -> None:
    """
    Асинхронно записывает категории в Excel файл
    Args:
        initial_data: Исходные данные категорий (список словарей)
        output_filename: Имя выходного файла Excel
        exclude_root_in_path: Исключать ли корневой элемент из пути
    """

    def _sanitize_sheet_name(name: str) -> str:
        """Удаляет недопустимые символы из названия листа Excel"""
        sanitized_name = re.sub(r'[\\/*?:\[\]]', '', name)
        sanitized_name = sanitized_name[:31]
        return sanitized_name

    def _get_categories_with_levels(
            item: dict,
            current_level: int = 0,
            parent_path: list[str] | None = None,
            last_category: str | None = None
    ) -> list[dict]:
        """Рекурсивно извлекает категории с уровнями вложенности"""
        categories: list[dict[str, str | int]] = []
        current_path = parent_path if parent_path is not None else []

        if not exclude_root_in_path or current_level > 0:
            current_path = current_path + [item['name']]

        if 'Категория' in item:
            item_categories = item['Категория']
            if isinstance(item_categories, list):
                for category in item_categories:
                    if category == "Категорий нет" and last_category is not None:
                        modified_path = current_path[:-1] if current_path else []
                        categories.append({
                            'Категория': last_category,
                            'Уровень': current_level,
                            'Путь': ' → '.join(modified_path) if modified_path else item['name']
                        })
                    else:
                        categories.append({
                            'Категория': category,
                            'Уровень': current_level,
                            'Путь': ' → '.join(current_path) if current_path else item['name']
                        })
                    if category != "Категорий нет":
                        last_category = category
            else:
                if item_categories == "Категорий нет" and last_category is not None:
                    modified_path = current_path[:-1] if current_path else []
                    categories.append({
                        'Категория': last_category,
                        'Уровень': current_level,
                        'Путь': ' → '.join(modified_path) if modified_path else item['name']
                    })
                else:
                    categories.append({
                        'Категория': item_categories,
                        'Уровень': current_level,
                        'Путь': ' → '.join(current_path) if current_path else item['name']
                    })
                if item_categories != "Категорий нет":
                    last_category = item_categories

        if 'subcategories' in item and item['subcategories']:
            for sub in item['subcategories']:
                categories.extend(_get_categories_with_levels(
                    sub,
                    current_level + 1,
                    current_path.copy(),
                    last_category
                ))
        return categories

    # Создаем Excel-файл в памяти
    wb = Workbook()
    wb.remove(wb.active)  # Удаляем дефолтный лист

    for section in initial_data:
        section_id = section.get('data_menu_id', 'N/A')
        section_name = _sanitize_sheet_name(f"{section['name']} (id {section_id})")

        all_categories: list[dict[str, str | int]] = []

        if 'subcategories' in section and section['subcategories']:
            for subcategory in section['subcategories']:
                all_categories.extend(_get_categories_with_levels(
                    subcategory,
                    current_level=1,
                    parent_path=[],
                    last_category=None
                ))
        elif 'Категория' in section:
            all_categories.extend(_get_categories_with_levels(
                section,
                current_level=0,
                parent_path=[],
                last_category=None
            ))

        # Создаем DataFrame
        df = pd.DataFrame(all_categories)
        if not df.empty:
            df = df[['Категория', 'Уровень', 'Путь']]

            # Создаем лист в рабочей книге
            ws = wb.create_sheet(title=section_name)

            # Записываем заголовки
            ws.append(['Категория', 'Уровень', 'Путь'])

            # Записываем данные
            for _, row in df.iterrows():
                ws.append([row['Категория'], row['Уровень'], row['Путь']])

            # Форматирование
            for row in ws.iter_rows(min_row=2, min_col=2, max_col=2):
                for cell in row:
                    cell.alignment = Alignment(horizontal='center', vertical='center')

            ws['B1'].alignment = Alignment(horizontal='center', vertical='center')
            ws.column_dimensions['A'].width = 30
            ws.column_dimensions['B'].width = 10
            ws.column_dimensions['C'].width = 50
        else:
            # Создаем пустой лист
            wb.create_sheet(title=section_name)

    # Сохраняем файл асинхронно
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    async with aiofiles.open(output_filename, 'wb') as f:
        await f.write(buffer.getvalue())


# Остальной код (Excel экспорт и другие вспомогательные функции) при необходимости можно добавить

if __name__ == "__main__":
    asyncio.run(run_wb_parser())
