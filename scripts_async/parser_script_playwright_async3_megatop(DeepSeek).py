import os
import re
import time
import asyncio
import pandas as pd
from functools import wraps
from openpyxl import Workbook
from openpyxl.styles import Alignment
import aiofiles
from io import BytesIO
from asyncio import to_thread
from colorama import Fore, Style, init
from playwright.async_api import (
    async_playwright,
    TimeoutError as PlaywrightTimeoutError,
    Page, Playwright, Browser, BrowserContext
)

init()

# Константы
TARGET_URL = "https://www.wildberries.by"
EXCLUDED_CATEGORIES = ['бренды', 'wibes', 'экспресс', 'акции', 'грузовая доставка']
TIME_WAIT = 1000
MAX_CONCURRENT_WORKERS = 10  # Количество параллельных воркеров


def timeit(func):
    @wraps(func)
    async def async_wrapper(*args, **kwargs):
        start_time = time.time()
        try:
            result = await func(*args, **kwargs)
            end_time = time.time()
            print(f'Функция {func.__name__} выполнилась за {end_time - start_time:.4f} сек.')
            return result
        except Exception as e:
            end_time = time.time()
            print(f'Функция {func.__name__} завершилась с ошибкой за {end_time - start_time:.4f} сек.')
            raise e

    def wrapper(*args, **kwargs):
        if asyncio.iscoroutinefunction(func):
            return async_wrapper(*args, **kwargs)
        else:
            start_time = time.time()
            result = func(*args, **kwargs)
            end_time = time.time()
            print(f'Синхронная функция {func.__name__} выполнилась за {end_time - start_time:.4f} сек.')
            return result

    return wrapper


async def process_categories_to_excel(
        initial_data: list[dict],
        output_filename: str = "categories.xlsx",
        exclude_root_in_path: bool = True
) -> None:
    """
    Обрабатывает категории и сохраняет их в Excel файл с уровнями вложенности
    Args:
        initial_data: Исходные данные категорий (список словарей)
        output_filename: Имя выходного файла Excel
        exclude_root_in_path: Исключать ли корневой элемент из пути
    """
    # Получаем уникальное имя файла
    unique_filename = await get_unique_filename(output_filename)

    # Вызываем функцию для записи данных в файл
    await _write_categories_to_excel(initial_data, unique_filename, exclude_root_in_path)


async def get_unique_filename(base_filename: str) -> str:
    """
    Асинхронно генерирует уникальное имя файла, если файл с таким именем уже существует.
    Args:
        base_filename: Базовое имя файла
    Returns:
        Уникальное имя файла
    """
    # Асинхронно проверяем существование файла
    if not await to_thread(os.path.exists, base_filename):
        return base_filename

    # Разделяем имя файла и расширение
    name, ext = os.path.splitext(base_filename)
    directory = os.path.dirname(base_filename) or '.'

    # Асинхронно получаем список файлов в директории
    existing_files = await to_thread(
        lambda: [f for f in os.listdir(directory)
                 if f.startswith(name) and f.endswith(ext)]
    )

    if not existing_files:
        return base_filename

    # Находим максимальный суффикс
    max_suffix = 0
    for f in existing_files:
        # Пытаемся извлечь суффикс из имени файла
        try:
            suffix_part = f[len(name):-len(ext)]
            # Обрабатываем случай с подчеркиванием (например, "file_1.xlsx")
            if '_' in suffix_part:
                suffix = int(suffix_part.split('_')[-1])
            else:
                suffix = int(suffix_part)
            if suffix > max_suffix:
                max_suffix = suffix
        except ValueError:
            continue

    # Генерируем новое имя файла
    new_filename = f"{name}_{max_suffix + 1}{ext}" if max_suffix > 0 else f"{name}_1{ext}"
    return new_filename


async def _write_categories_to_excel(
        initial_data: list[dict],
        output_filename: str,
        exclude_root_in_path: bool
) -> None:
    """
    Асинхронно записывает категории в Excel файл
    Args:
        initial_data: Исходные данные категорий (список словарей)
        output_filename: Имя выходного файла Excel
        exclude_root_in_path: Исключать ли корневой элемент из пути
    """

    def _sanitize_sheet_name(name: str) -> str:
        """Удаляет недопустимые символы из названия листа Excel"""
        sanitized_name = re.sub(r'[\\/*?:\[\]]', '', name)
        sanitized_name = sanitized_name[:31]
        return sanitized_name

    def _get_categories_with_levels(
            item: dict,
            current_level: int = 0,
            parent_path: list[str] | None = None,
            last_category: str | None = None
    ) -> list[dict]:
        """Рекурсивно извлекает категории с уровнями вложенности"""
        categories: list[dict[str, str | int]] = []
        current_path = parent_path if parent_path is not None else []

        if not exclude_root_in_path or current_level > 0:
            current_path = current_path + [item['name']]

        if 'Категория' in item:
            item_categories = item['Категория']
            if isinstance(item_categories, list):
                for category in item_categories:
                    if category == "Категорий нет" and last_category is not None:
                        modified_path = current_path[:-1] if current_path else []
                        categories.append({
                            'Категория': last_category,
                            'Уровень': current_level,
                            'Путь': ' → '.join(modified_path) if modified_path else item['name']
                        })
                    else:
                        categories.append({
                            'Категория': category,
                            'Уровень': current_level,
                            'Путь': ' → '.join(current_path) if current_path else item['name']
                        })
                    if category != "Категорий нет":
                        last_category = category
            else:
                if item_categories == "Категорий нет" and last_category is not None:
                    modified_path = current_path[:-1] if current_path else []
                    categories.append({
                        'Категория': last_category,
                        'Уровень': current_level,
                        'Путь': ' → '.join(modified_path) if modified_path else item['name']
                    })
                else:
                    categories.append({
                        'Категория': item_categories,
                        'Уровень': current_level,
                        'Путь': ' → '.join(current_path) if current_path else item['name']
                    })
                if item_categories != "Категорий нет":
                    last_category = item_categories

        if 'subcategories' in item and item['subcategories']:
            for sub in item['subcategories']:
                categories.extend(_get_categories_with_levels(
                    sub,
                    current_level + 1,
                    current_path.copy(),
                    last_category
                ))
        return categories

    # Создаем Excel-файл в памяти
    wb = Workbook()
    wb.remove(wb.active)  # Удаляем дефолтный лист

    for section in initial_data:
        section_id = section.get('data_menu_id', 'N/A')
        section_name = _sanitize_sheet_name(f"{section['name']} (id {section_id})")

        all_categories: list[dict[str, str | int]] = []

        if 'subcategories' in section and section['subcategories']:
            for subcategory in section['subcategories']:
                all_categories.extend(_get_categories_with_levels(
                    subcategory,
                    current_level=1,
                    parent_path=[],
                    last_category=None
                ))
        elif 'Категория' in section:
            all_categories.extend(_get_categories_with_levels(
                section,
                current_level=0,
                parent_path=[],
                last_category=None
            ))

        # Создаем DataFrame
        df = pd.DataFrame(all_categories)
        if not df.empty:
            df = df[['Категория', 'Уровень', 'Путь']]

            # Создаем лист в рабочей книге
            ws = wb.create_sheet(title=section_name)

            # Записываем заголовки
            ws.append(['Категория', 'Уровень', 'Путь'])

            # Записываем данные
            for _, row in df.iterrows():
                ws.append([row['Категория'], row['Уровень'], row['Путь']])

            # Форматирование
            for row in ws.iter_rows(min_row=2, min_col=2, max_col=2):
                for cell in row:
                    cell.alignment = Alignment(horizontal='center', vertical='center')

            ws['B1'].alignment = Alignment(horizontal='center', vertical='center')
            ws.column_dimensions['A'].width = 30
            ws.column_dimensions['B'].width = 10
            ws.column_dimensions['C'].width = 50
        else:
            # Создаем пустой лист
            wb.create_sheet(title=section_name)

    # Сохраняем файл асинхронно
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    async with aiofiles.open(output_filename, 'wb') as f:
        await f.write(buffer.getvalue())


async def category_worker(queue: asyncio.Queue, context: BrowserContext, results: list):
    """Воркер для обработки категорий из очереди"""
    while True:
        try:
            # Получаем задание из очереди
            task = await queue.get()
            if task is None:  # Сигнал завершения
                queue.task_done()
                break

            category, level, parent_path = task
            print(f'{level * "-" + " " if level != 1 else ""}{category["name"]}')

            # Обрабатываем категорию
            processed = await process_category(context, category, level, parent_path)

            if processed:
                results.append(processed)

        except asyncio.CancelledError:
            break
        except Exception as e:
            print(f"{Fore.RED}Ошибка в воркере: {str(e)}{Style.RESET_ALL}")
        finally:
            queue.task_done()


async def process_category(context: BrowserContext, category: dict, level: int, parent_path: list) -> dict:
    """Обработка одной категории"""
    MAX_ATTEMPTS = 3
    attempt = 0

    while attempt < MAX_ATTEMPTS:
        attempt += 1
        page = await context.new_page()
        try:
            await page.route("**/*", route_handler)
            await page.goto(
                f"{TARGET_URL}{category['url']}",
                timeout=120000,
                wait_until="networkidle"
            )

            # Проверяем разные типы меню
            menu_subcategory = page.locator('ul.menu-category__subcategory')
            menu_list = page.locator('ul.menu-category__list')
            category_filter = page.locator("div.dropdown-filter:has-text('Категория'):visible")
            btm_burger = page.locator('button.dropdown-filter__btn--burger > div.dropdown-filter__btn-name').first

            if await menu_subcategory.count() > 0:
                subcategories = await extract_menu_items(page, 'subcategory', category, level, parent_path)
            elif await menu_list.count() > 0:
                subcategories = await extract_menu_items(page, 'list', category, level, parent_path)
            elif await category_filter.count() > 0:
                await category_filter.hover()
                await load_and_collect_categories(page, category, level)
            elif await btm_burger.count() > 0:
                await btm_burger.hover()
                await load_and_collect_subcategories(page, context, category, level)
            else:
                category['Категория'] = 'Категорий нет'

            return category

        except Exception as e:
            print(f"{Fore.RED}Ошибка обработки {category['name']}: {str(e)}{Style.RESET_ALL}")
        finally:
            await page.close()

    return None


async def extract_menu_items(page: Page, menu_type: str, category: dict, level: int, parent_path: list) -> list:
    """Извлекает элементы меню"""
    selector = {
        'subcategory': ('li.menu-category__subcategory-item', 'a.menu-category__subcategory-link'),
        'list': ('li.menu-category__item', 'a.menu-category__link')
    }[menu_type]

    items = await page.locator(selector[0]).all()
    subcategories = []

    for item in items:
        if menu_type == 'list' and await item.locator('p.menu-category__item').count() > 0:
            continue

        link = item.locator(selector[1])
        if await link.count() > 0:
            subcategories.append({
                'name': (await link.inner_text()).strip(),
                'url': await link.get_attribute('href'),
                'parent': parent_path + [category['name']],
                'level': level + 1,
            })

    return subcategories


async def run_wb_parser():
    result = await parse_all_categories()
    await process_categories_to_excel(result)


@timeit
async def parse_all_categories() -> list:
    """Основная функция парсинга с использованием очереди"""
    async with async_playwright() as p:
        browser, context = await create_browser_session(p)
        results = []
        queue = asyncio.Queue()

        try:
            # 1. Загружаем основные категории
            main_categories = await load_main_categories(context)

            # 2. Создаем и запускаем воркеры
            workers = [
                asyncio.create_task(category_worker(queue, context, results))
                for _ in range(MAX_CONCURRENT_WORKERS)
            ]

            # 3. Добавляем начальные задачи в очередь
            for category in main_categories:
                await queue.put((category, 1, []))

            # 4. Ждем завершения обработки
            await queue.join()

            # 5. Отправляем сигналы завершения воркерам
            for _ in range(MAX_CONCURRENT_WORKERS):
                await queue.put(None)

            # 6. Ждем завершения воркеров
            await asyncio.gather(*workers, return_exceptions=True)

            return results

        finally:
            await context.close()
            await browser.close()


async def create_browser_session(p: Playwright) -> tuple[Browser, BrowserContext]:
    """
    Настраивает и возвращает экземпляр браузера, контекст и страницу Playwright.

    Args:
        p (Playwright): Экземпляр Playwright.

    Returns:
        tuple[Browser, BrowserContext, Page]: Экземпляр браузера, контекст и страница.
    """
    # Запускаем браузер с дополнительными параметрами
    browser = await p.chromium.launch(
        headless=True,
        # headless=False,
        timeout=60000,  # Увеличиваем таймаут запуска браузера
        slow_mo=0,  # Добавляем задержку между действиями (мс)
        args=[
            '--disable-blink-features=AutomationControlled',
            '--disable-infobars',
            '--start-maximized',
            '--no-sandbox',
            '--disable-setuid-sandbox',
            '--disable-dev-shm-usage',
            '--disable-web-security',
            '--disable-notifications'
        ]
    )

    # Создание контекста (окна браузера с настройками)
    context = await browser.new_context(
        viewport={'width': 1280, 'height': 720},
        user_agent='Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36',
        # Дополнительные настройки защиты от обнаружения
        bypass_csp=True,
        java_script_enabled=True
    )

    return browser, context


async def route_handler(route, request):
    """Блокировка "лишних" ресурсов (ускорение загрузки)"""
    try:
        # blocked_types = ["image", "font", "stylesheet", "media", "other"]
        blocked_types = ["image", "media", "other"]
        if request.resource_type in blocked_types:
            await route.abort()
        else:
            await route.continue_()
    except Exception as e:
        print(f"Error in route handler for {request.url}: {str(e)}")
        await route.continue_()  # В случае ошибки лучше пропустить запрос


@timeit
async def load_main_categories(context: BrowserContext) -> list:
    """
    Загружает основные категории с веб-страницы, используя Playwright.

    Эта функция открывает новую страницу в браузере, переходит по указанному URL,
    ожидает загрузки меню и извлекает основные категории, исключая указанные.

    Args:
        context (BrowserContext): Контекст браузера Playwright для открытия новой страницы.

    Returns:
        list: Список словарей, содержащих информацию о каждой категории,
              включая название, URL и идентификатор меню.
    """
    result = []
    # Открытие новой вкладки
    page = await context.new_page()
    await page.route("**/*", route_handler)

    try:
        # Увеличиваем таймаут навигации и отключаем ожидание полной загрузки
        await page.goto(
            f"{TARGET_URL}",
            timeout=120000,  # 2 минуты на загрузку
            wait_until="networkidle"  # Ждём, пока не закончатся сетевые запросы
        )

        # Ожидаем появления кнопки меню и кликаем по ней
        button_burger = page.locator('button.nav-element__burger.j-menu-burger-btn').first
        if button_burger:
            await button_burger.click()

        # Ждем появления меню с категориями
        await page.wait_for_selector(
            'ul.menu-burger__main-list a.menu-burger__main-list-link:has-text("Бренды"):visible')

        # Получаем категории
        main_categories = await page.locator('ul.menu-burger__main-list > li.menu-burger__main-list-item').all()

        for category in main_categories:
            category_link = category.locator('a.menu-burger__main-list-link')
            if not category_link:
                continue

            name_category = (await category_link.inner_text()).strip()
            if name_category.lower() in EXCLUDED_CATEGORIES:
                continue

            result.append({
                'name': (await category_link.inner_text()).strip(),
                'url': await category_link.get_attribute('href'),
                'data_menu_id': await category.get_attribute('data-menu-id'),
                'subcategories': []
            })

    except PlaywrightTimeoutError:
        print(f"Таймаут ожидания элемента")
    except Exception as e:
        print(f"{Fore.RED}Обработка ошибок: {str(e)}{Style.RESET_ALL}")
    finally:
        await page.close()

    return result


async def load_and_collect_categories(
        page: Page,
        category: dict,
        level: int) -> None:
    """
        Загружает и собирает категории с веб-страницы, используя Playwright.

        Эта функция пытается найти и нажать кнопку "Показать все", чтобы загрузить все категории.
        Затем она собирает все видимые категории и добавляет их в переданный словарь категории.

        Args:
            page (Page): Страница Playwright, на которой происходит поиск и взаимодействие с элементами.
            category (dict): Словарь, представляющий категорию, в которую будут добавлены собранные подкатегории.
            level (int): Уровень вложенности категории, используется для форматирования вывода.
    """
    await page.wait_for_timeout(100)
    show_all_button = page.locator("button.filter__show-all:has-text('Показать все'):visible").first
    if show_all_button and await show_all_button.is_visible():
        await show_all_button.click()

        current_items = await page.locator('li.filter__item:visible').all()
        while True:
            if not current_items:
                await page.wait_for_timeout(500)
                current_items = await page.locator('li.filter__item:visible').all()
                continue

            await current_items[-1].hover()
            await page.wait_for_timeout(100)

            new_items = await page.locator('li.filter__item:visible').all()
            if len(new_items) == len(current_items):
                break

            current_items = new_items
    else:
        current_items = await page.locator('li.filter__item:visible').all()

    result = []
    for item in current_items:
        link = item.locator('span.checkbox-with-text__text')
        if link:
            result.append((await link.inner_text()).strip())

    # print(f'{Fore.YELLOW}{level * "-" + "-"} Категорий {len(result)}, [Родитель: {category["name"]}], {level}{Style.RESET_ALL}')
    category['Категория'] = result


async def load_and_collect_subcategories(
        page: Page,
        context: BrowserContext,
        category: dict,
        level: int,
        max_concurrent_tasks: int = 2
) -> None:
    """
        Загружает и собирает подкатегории с веб-страницы, используя Playwright.

        Эта функция собирает все видимые подкатегории для заданной категории и добавляет их в словарь категории.
        Для каждой подкатегории также загружаются её подкатегории рекурсивно.

        Args:
            page (Page): Страница Playwright, на которой происходит поиск и взаимодействие с элементами.
            context (BrowserContext): Контекст браузера Playwright, используемый для открытия новых страниц.
            category (dict): Словарь, представляющий категорию, в которую будут добавлены собранные подкатегории.
            level (int): Уровень вложенности категории, используется для форматирования вывода.
            max_concurrent_tasks (int): Указывает конкретное число ограничения.
    """
    result = []
    await page.wait_for_timeout(100)
    all_items = await page.locator('ul.filter-category__list > li.filter-category__item').all()
    for item in all_items:
        # Извлекаем ссылки
        link = item.locator('a.filter-category__link').first
        if link:
            parent = category['parent'].copy() if category.get('parent') else []
            parent.append(category['name'])
            result.append({
                'name': (await link.inner_text()).strip(),
                'url': await link.get_attribute('href'),
                'parent': parent,
                'level': level,
            })
    # print(f"{Fore.BLUE}{[i['name'] for i in result]}{Style.RESET_ALL}")
    category['subcategories'] = []

    if result:
        # print(f'Получаю подкатегории {level}')
        sub_sem = asyncio.Semaphore(max_concurrent_tasks)
        tasks = [load_subcategories(item, context, sub_sem, level + 1, False) for item in result]
        category['subcategories'] = await asyncio.gather(*tasks, return_exceptions=False)


# Остальные функции (create_browser_session, route_handler, load_main_categories,
# load_and_collect_categories, load_and_collect_subcategories,
# process_categories_to_excel, get_unique_filename, _write_categories_to_excel)
# остаются без изменений из вашего исходного кода

if __name__ == "__main__":
    asyncio.run(run_wb_parser())